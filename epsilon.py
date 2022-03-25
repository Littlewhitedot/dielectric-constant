# 作者Hai
import tkinter
from scipy import special
import math
import openpyxl
import sys

# 主窗体
top = tkinter.Tk(className='Epsilon calculation', )

# 定义窗体大小及位置
width = 650
height = 240
screenwidth = top.winfo_screenwidth()
screenheight = top.winfo_screenheight()
alignstr = '%dx%d+%d+%d' % (width, height, (screenwidth-width)/2, (screenheight-height)/2)
top.geometry(alignstr)

# 加上标签
label = tkinter.Label(top)
label['text'] = 'made by Z in TianJi, GuangZhou'
label.place(y=210, x=225)


# 计算k值的函数
def calculation(thickness, diameter, frequency):
    # parameters
    T = thickness
    t = T * 1e-3
    D = diameter
    d = D * 1e-3
    f = frequency
    f0 = f * 1e9
    a = d / 2
    c = 3e8
    # calculate
    lambda0 = c / f0
    k_c0 = (2 * math.pi / lambda0) * math.sqrt((lambda0 / 2 / t) * (lambda0 / 2 / t) - 1)
    k = range(10, 800, 1)
    ys = []
    for k_c in k:
        y = special.jv(0, k_c * a) * k_c * special.kv(1, k_c0 * a) + special.jv(1, k_c * a) * k_c0 * special.kv(0,
                                                                                                                k_c0 * a)
        y = abs(y)
        ys.append(y)
        # print(y)
    N = 0
    nmin = 0
    Y = ys[0]
    for i in ys:
        N = N + 1
        if i < Y:
            Y = i
            nmin = N - 1
    k_ci = k[nmin]
    # finally
    eleconstant = (lambda0 / 2 / math.pi) * (lambda0 / 2 / math.pi) * (k_ci * k_ci + k_c0 * k_c0) + 1
    # print(eleconstant)
    return round(eleconstant, 5)


# 获取输入框的值
def get_value(value):
    v = value.get()
    return float(v)


# 定义按钮“计算”的功能
def cal_click():
    thickness = get_value(first_input)
    diameter = get_value(second_input)
    frequence = get_value(third_input)
    k = calculation(thickness, diameter, frequence)
    print('通过按钮获取的输入值为： 厚度t=%.2f 直径d=%.2f 频率f=%.4f' % (thickness, diameter, frequence))
    print('计算得到的相对介电常数为： ', k)
    K = str(k)
    outcome['text'] = '相对介电常数为 k = '+K


# 下拉菜单
def helpme():
    helpwindow = tkinter.Tk(className='软件介绍')
    allindex = '%dx%d+%d+%d' % (600, 190, (screenwidth - width) / 2, (screenheight - height) / 2)
    helpwindow.geometry(allindex)
    helplabel = tkinter.Label(helpwindow, font=('宋体', 15), text='  \n该软件的功能为计算微波介电陶瓷的相对介电常数。\n'
                                                                '所用方法为平行板谐振法。\n'
                                                                '将陶瓷粉体压制为圆柱体并完成烧结，\n'
                                                                '利用卡尺精确测出其厚度与直径，利用网络分析仪测出其谐振频率。\n'
                                                                '将此三参数对应输入后便可计算其相对介电常数。\n'
                                                                '\n'
                                                                '注意：使用各功能前请仔细阅读各功能的操作方法！注意单位！')
    helplabel.pack()


def instruction_single():
    helpwindow = tkinter.Tk(className='单值计算使用说明')
    allindex = '%dx%d+%d+%d' % (600, 200, (screenwidth - width) / 2, (screenheight - height) / 2)
    helpwindow.geometry(allindex)
    helplabel = tkinter.Label(helpwindow, font=('宋体', 15), text='\n\n该按钮的作用为计算一组数据，无存储功能\n\n'
                                                                '1.在指定位置输入样品的厚度、直径及谐振频率\n'
                                                                '2.点击“单值计算”(浅珊瑚色)按钮\n'
                                                                '3.计算结果显示在输出框内')
    helplabel.pack()


def instruction_save():
    helpwindow = tkinter.Tk(className='单值存储')
    allindex = '%dx%d+%d+%d' % (730, 230, (screenwidth - width) / 2, (screenheight - height) / 2)
    helpwindow.geometry(allindex)
    helplabel = tkinter.Label(helpwindow, font=('宋体', 15), text='\n该按钮的作用为计算一组数据，并将数据按指定顺序存入指定的excel中\n\n'
                                                                '1.在指定位置输入样品的厚度、直径及谐振频率\n'
                                                                '2.在样品编号处输入其序号(正整数)\n'
                                                                '3.点击“单值存储”(天蓝色)按钮,点击按钮前请关闭excel文件\n'
                                                                '4.运行结果显示在输出框内\n\n'
                                                                '注意：使用该功能前需要在该程序相同的路径(文件夹)下创建一个名\n'
                                                                '   为“database.xlsx”的excel文件，否则程序无法正确运行，数据存入该文件中\n')
    helplabel.pack()


def instruction_multiple():
    helpwindow = tkinter.Tk(className='连续计算')
    allindex = '%dx%d+%d+%d' % (730, 260, (screenwidth - width) / 2, (screenheight - height) / 2)
    helpwindow.geometry(allindex)
    helplabel = tkinter.Label(helpwindow, font=('宋体', 15), text='\n该按钮的作用为批量计算多组数据，并将数据按对应顺序存入指定的excel中\n\n'
                                                                '1.在该程序相同的路径(文件夹)下创建一个名为“database.xlsx”的excel文件\n'
                                                                '2.将厚度、直径、频率写入各自对应的位置，\n'
                                                                '  具体为：excel的B列填写厚度数据，C列填写直径数据，D列填写频率数据\n'
                                                                '3.点击“连续计算”按钮，点击该按钮前请关闭excel文件\n'
                                                                '4.运行结果显示在输出框内，数据存入excel中，K值放于E列\n\n'
                                                                '注意：使用该功能前需要在该程序相同的路径(文件夹)下创建一个名\n'
                                                                '     为“database.xlsx”的excel文件，并且按照规定方式写入数据，\n'
                                                                '     否则程序无法正确运行，数据将存入该文件中')
    helplabel.pack()


# 下拉菜单
def otherthings():
    otherwindow = tkinter.Tk(className='其他说明')
    allindex = '%dx%d+%d+%d' % (410, 95, (screenwidth - width) / 2, (screenheight - height) / 2)
    otherwindow.geometry(allindex)
    otherlabel = tkinter.Label(otherwindow, font=('宋体', 15), text='制作该软件所用的工具为公开的免费软件\n'
                                                                  '编写该软件所用的语言为开源语言\n'
                                                                  '该软件不涉及任何盗版侵权问题\n'
                                                                  '请放心使用！')
    otherlabel.pack()


# 定义按钮“保存”的功能
def save_click():
    sequence, thickness, diameter, frequence, epslion = save_action()
    # print(thickness,  diameter,  frequence,  int(sequence))
    K = str(epslion)
    outcome['text'] = '相对介电常数为 k = '+K+'\n已存储'
    print('%d号样品存储的数据为：厚度t=%.2f 直径d=%.2f 频率f=%.4f \n保存成功！数据存储于database.xlsx' % (sequence, thickness, diameter, frequence))


# 实现保存数据入excel
def save_action():
    thickness = get_value(first_input)
    diameter = get_value(second_input)
    frequence = get_value(third_input)
    sequence = int(get_value(save_input))
    epslion = calculation(thickness, diameter, frequence)
    workbook = openpyxl.load_workbook('database.xlsx')
    sheet = workbook.active
    sheet['A1'] = '序号'
    sheet['B1'] = '厚度(mm)'
    sheet['C1'] = '直径(mm)'
    sheet['D1'] = '频率(GHz)'
    sheet['E1'] = 'K'
    sheet.cell(sequence + 1, 1).value = sequence
    sheet.cell(sequence + 1, 2).value = thickness
    sheet.cell(sequence + 1, 3).value = diameter
    sheet.cell(sequence + 1, 4).value = frequence
    sheet.cell(sequence + 1, 5).value = epslion
    workbook.save('database.xlsx')
    return sequence, thickness, diameter, frequence, epslion


# 定义按钮‘连续计算’的功能
def continue_click():
    numbers = continue_action()
    workbook = openpyxl.load_workbook('database.xlsx')
    sheet = workbook.active
    alldata = sheet.rows
    for row in alldata:
        print(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)
    outcome['text'] = '共计算{0}组数据，并存入excel'.format(numbers)


# 实现连续计算
def continue_action():
    workbook = openpyxl.load_workbook('database.xlsx')
    sheet = workbook.active
    alldata = sheet.rows
    numbers = 0
    for row in alldata:
        if isinstance(row[1].value, (int, float)) and isinstance(row[2].value, (int, float)) and isinstance(row[3].value, (int, float)):
            thickness = float(row[1].value)
            diameter = float(row[2].value)
            frequence = float(row[3].value)
            epslion = calculation(thickness, diameter, frequence)
            rownum = row[1].row
            sheet.cell(rownum, 5).value = epslion
            numbers = numbers+1
    workbook.save('database.xlsx')
    return numbers


# 创建一个输入框,并设置尺寸
first_input = tkinter.Entry(top, width=8)
first_text = tkinter.DoubleVar()
first_text.set(4.56)
first_input['textvariable'] = first_text
first_input.place(x=200, y=20)
label1 = tkinter.Label(top, width=25, text='请输入圆柱体厚度(单位mm)： ')
label1.place(x=14, y=20)

# 创建一个输入框,并设置尺寸
second_input = tkinter.Entry(top, width=8)
second_text = tkinter.DoubleVar()
second_text.set(8.56)
second_input['textvariable'] = second_text
second_input.place(x=200, y=60)
label2 = tkinter.Label(top, width=25, text='请输入圆柱体直径(单位mm)： ')
label2.place(x=14, y=60)

# 创建一个输入框,并设置尺寸
third_input = tkinter.Entry(top, width=8)
third_text = tkinter.DoubleVar()
third_text.set(12.85)
third_input['textvariable'] = third_text
third_input.place(x=200, y=100)
label3 = tkinter.Label(top, width=25, text='请输入谐振频率(单位GHz)： ')
label3.place(x=14, y=100)

# 创建一个输入框
save_input = tkinter.Entry(top, width=5)
save_text = tkinter.DoubleVar()
save_text.set(1)
save_input['textvariable'] = save_text
save_input.place(x=420, y=170)
label4 = tkinter.Label(top, width=32, text='请输入该样品编号(正整数，如1、2、3等)：')
label4.place(x=185, y=170)

# 创建一个存储按钮
save_button = tkinter.Button(top, bg='skyblue', height=2, width=10, text='单值存储', command=lambda: save_click())
save_button.place(x=465, y=160)

# 创建一个查询结果的按钮
result_button = tkinter.Button(top, bg='lightcoral', height=3, width=12, command=lambda: cal_click(), text='单值计算')
result_button.place(x=50, y=150)


# 创建一个连续计算的按钮
continue_button = tkinter.Button(top, bg='orange', height=5, width=12, text='连续计算', command=lambda: continue_click())
continue_button.place(x=550, y=20)


# 创建一个显示结果的标签
outcome = tkinter.Label(top, bg='Tomato', fg='white', font=('华文行楷', 15), width=28, height=4)
outcome['text'] = '请输入数值进行计算'
outcome.place(x=280, y=30)
# print('不通过按钮获取的输入框值为： ', get_value(first_input), get_value(second_input), get_value(third_input))


menubar = tkinter.Menu(top)
# 创建下拉菜单1
helpmenu = tkinter.Menu(menubar, tearoff=0)
helpmenu.add_command(label='软件介绍', command=helpme)
helpmenu.add_command(label='单值计算使用说明', command=instruction_single)
helpmenu.add_command(label='单值存储使用说明', command=instruction_save)
helpmenu.add_command(label='连续计算使用说明', command=instruction_multiple)
menubar.add_cascade(label='Help', menu=helpmenu)

# 创建下拉菜单2
othermenu = tkinter.Menu(menubar, tearoff=0)
othermenu.add_command(label='其他说明', command=otherthings)
menubar.add_cascade(label='Other', menu=othermenu)


# 进入消息循环体
top.config(menu=menubar)
top.mainloop()

# 测试
print('冰冻三尺非一日之寒！')
